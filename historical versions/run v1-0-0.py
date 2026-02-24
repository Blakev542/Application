import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tkinter import Tk, filedialog
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter 
import customtkinter as ctk


class ExcelCombinerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Column Combiner")
        self.files = []
        self.allcolumns = []
        self.columns = []
        self.selected = []
        self.df = pd.DataFrame()
        self.generated_df = pd.DataFrame()
        self.ALIASES = {
            "eqid": "eq_id",
            "student_id": "eq_id",
            "studentid": "eq_id",
        }
        self.view = []
        # UI Layout
        self.build_ui(

        ) 

    # ---------------- UI ----------------
    def build_ui(self):
        self.frame = tk.Frame(self.root, padx=5, pady=5)
        self.frame.pack(fill="both", expand=True)

        # File list
        tk.Label(self.frame, text="Selected Excel Files").grid(row=0, column=0, sticky="w" )
        self.file_listbox = tk.Listbox(self.frame, width=60, height=15)
        self.file_listbox.grid(row=1, column=0, columnspan=2, sticky="w",padx= 5 , pady=5)

        tk.Button(self.frame, text="Add Excel Files", command=self.add_files).grid(row=2, column=0, pady=5,sticky="w")
        tk.Button(self.frame, text="Clear Files", command=self.clear_files).grid(row=2, column=1,sticky="w", pady=5)

        # Column selector
        tk.Label(self.frame, text="Select Columns to Combine").grid(row=3, column=0, sticky="w")
        self.column_search_bar = tk.Entry(self.frame, text= "")
        self.column_search_bar.grid(row=3, column= 1, sticky = "w", pady=5)
        self.column_search_bar.bind("<Key>", self.show_columns_listbox)


        self.column_listbox = tk.Listbox(self.frame, selectmode="multiple", width=60, height=20)
        self.column_listbox.grid(row=4, column=0, pady=5, columnspan=2,padx= 5)

        tk.Button(self.frame, text="Load Columns", command=self.load_columns).grid(row=5, column=0, pady=5,sticky="w")
        tk.Button(self.frame, text="Search", command=self.show_columns_listbox).grid(row=5, column=1, pady=5,sticky="w")
        self.frame.grid_propagate(False)


        tk.Label(self.frame, text="Data Preview").grid(row=0, column=3, sticky="w", pady =5)

        self.tree = ttk.Treeview(self.frame)
        self.tree.grid(row=1, column=3, sticky= "nsew",rowspan=4, columnspan=3, pady=5)
       
        self.frame.grid_columnconfigure(3, weight=1)
        self.frame.grid_rowconfigure(1, weight=1)


        tk.Button(self.frame, text="Add Column", command=self.show_columns).grid(row=5, column=3, pady=5)
        tk.Button(self.frame, text="Combine & Export", command=self.combine_and_export).grid(row=5, column=4,pady=5)
        tk.Button(self.frame, text="Preview Selected Columns",).grid(row=5, column=5, pady=5)

    # ---------------- File Handling ----------------
    def add_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx *.xls")])
        for f in files:
            if f not in self.files:
                self.files.append(f)
                self.file_listbox.insert(tk.END, f)
      
    def clear_cache(self):
        for col in self.tree.get_children():
            self.tree.delete(col)
        self.columns = []
        self.allcolumns.clear()

                    
                    
    def normalize_key(self,key):
        key = self.ALIASES.get(key.replace(" ", "_").lower(), key.replace(" ", "_").lower())
        return key 
   
    def find_best_key(self, df, candidates):
        df_candidates = []
        df_final_candidates = []
        for c in df.columns:
            if df[c].nunique() > len(df) * 0.9:  # mostly unique
                
                df_candidates.append(c)
        for candidate in df_candidates:
            if candidate in candidates:
                df_final_candidates.append(candidate)   
            
        return df_final_candidates
    def sort_candidates_by_uniqueness(self, df, candidates):
        scored = []

        for c in candidates:
            if c in df.columns:
                score = df[c].nunique(dropna=True)
                scored.append((c, score))

        # highest uniqueness first
        scored.sort(key=lambda x: x[1], reverse=True)

        return [c for c, _ in scored]
    def find_keys(self):
        default_key = "eq_id"
        columns = []
        countmax = 0
        candidates = []
        for column in self.allcolumns:
            columns.append(self.normalize_key(column))
        for i in columns:
            if columns.count(i) > 1:
                candidates.append(i)
        for i in candidates:
            if candidates.count(i)>countmax:
                countmax = candidates.count(i)
                default_key = i
        candidates = list(set(candidates))
        self.primary_key = default_key
        return default_key, candidates
    def aggregate_to_lists(self, df, key):
      
        df = df.copy()

        agg_dict = {}
        #parent_lookup_email = parents.groupby("EQ_Id")["Email_Address"].apply(list).to_dict()
        if key not in df.columns: return
        for column in df.columns:
            lists_for_column = df.groupby(key)[column].apply(list).to_dict()
            inner_dict_keys_list = list(lists_for_column.keys())
            for inner_dict_key in inner_dict_keys_list:
                combined_string = ""
             
                if any(datapoint for datapoint in lists_for_column[inner_dict_key]) != any(datapoint for datapoint in lists_for_column[inner_dict_key]) :
                    for data in lists_for_column[inner_dict_key]:
                        if combined_string == "":
                            combined_string = str(data)
                        else:
                            combined_string +=(", "+ str(data) )
                          
                else: 
                    combined_string =  str(lists_for_column[inner_dict_key][0])
                lists_for_column[inner_dict_key] = combined_string




                
                agg_dict[column] = lists_for_column
       
        df = pd.DataFrame(agg_dict)
       
        return df 
        # Keep key as first value (not list)
       
        


    def build_df(self):
        default_key = "eq_id"
        filesnotmerged = self.files.copy()
        default_key, candidates = self.find_keys()
        maxloops = 20
        loops = 0
        self.df = pd.DataFrame()

        while(len(filesnotmerged) > 0):
            loops += 1
     
            if loops >= maxloops:
                print("exceeded maximum loops")
                return self.df
                
            for file in filesnotmerged[:]:
                filename = file.rsplit("/", 1)[-1]
                df = pd.read_excel(file)
                df = self.clean_file(df)
                df = df.rename(columns=lambda c: self.normalize_key(c))
                
                df_candidates = []
                for i in candidates:
                    if i in self.find_best_key(df, candidates):
                        df_candidates.append(i)

                if default_key not in df_candidates and len(df_candidates) != 0:
             
                    default_key = df_candidates[0]
                #if len(df_candidates) == 0:
                #    df = self.aggregate_to_lists(df,default_key)

                if self.df.empty and default_key in df.columns:

                    self.df = df

                    if file in filesnotmerged:filesnotmerged.remove(file)
                    continue
                elif default_key in self.df.columns and default_key in df.columns and default_key in df_candidates:

                    self.df = df.merge(self.df, how="outer", on=default_key, suffixes=("_"+ filename,""))
                    
                    if file in filesnotmerged:filesnotmerged.remove(file)
                    continue
                else:
                    candidates = self.sort_candidates_by_uniqueness(df, candidates)
                    for candidate_key in candidates:
                        if candidate_key in self.df.columns and candidate_key in df.columns:
                            #if candidate_key not in df_candidates:
                                #df = self.aggregate_to_lists(df,candidate_key)
                            print("\ndefault key" , candidate_key,"\n\n\n\n\n")
                           
                            self.df = df.merge(self.df, how="outer", on=candidate_key, suffixes=("_"+ filename,""))

                            if file in filesnotmerged:filesnotmerged.remove(file)
                            break  
        self.columns = self.df.columns
        return self.df     
    def show_columns(self):
        self.selected = [self.column_listbox.get(i) for i in self.column_listbox.curselection()]
        for col in self.tree.get_children():
            self.tree.delete(col)
        self.generated_df = self.df.copy()
        self.view.extend(self.selected)
        self.generated_df.drop(columns=[c for c in self.df.columns if c not in self.view], inplace=True)
        self.generated_df = self.generated_df.drop_duplicates()
        self.generated_df = self.generated_df[self.view]
        self.show_columns_listbox()
        self.tree["columns"] = self.view
        self.tree["show"] = "headings"
     
        for col in self.view:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, stretch=False)
        for _, row in self.generated_df.iterrows():
            self.tree.insert("", tk.END, values=row.tolist())
                            
    def show_columns_listbox(self):
    
        self.columns = [c for c in self.columns if c not in self.view]
        string = self.column_search_bar.get()
        if string is not None and string != "":
            string = string.lower()
            columns = [c for c in self.columns if string in c.lower()]
        else:
            columns = self.columns

        for col in columns:
            self.column_listbox.insert(tk.END, col)




    def clear_files(self):
        self.files.clear()
        self.file_listbox.delete(0, tk.END)
        self.column_listbox.delete(0, tk.END)
        self.columns.clear()

    # ---------------- Column Loading ----------------
    def load_columns(self):
        self.clear_cache()
        
        if not self.files:
            messagebox.showerror("Error", "No files selected")
            return
        
        
        
        for file in self.files:
            df = pd.read_excel(file)
     
            df = self.clean_file(df)
         
            for col in df.columns:
                col = col.replace(" ", "_").lower()
                self.allcolumns.append(str(col))
  
        
        self.build_df()
        self.show_columns()


    def clean_file(self, df):   

        
        first_clean_index = 0
        for i, row in df.iterrows():
            if not any(str(cell).startswith("Unnamed") for cell in row) and not any(pd.isna(cell) for cell in row):
                first_clean_index = i
                if i > 10:
                    return df
                df.columns = df.iloc[first_clean_index]



                df = df.iloc[first_clean_index+1:].reset_index(drop=True)

                return df
        
        return df 
            
           

    # ---------------- Combine + Export ----------------
    def combine_and_export(self):

        #selected = [self.column_listbox.get(i) for i in self.column_listbox.curselection()]
        
        #if not selected:
        #   messagebox.showerror("Error", "No columns selected")
        #   return

        output_folder = "Output"
        if not output_folder:
            return
          
        output_path = "Output/output.xlsx"

        # Write with XlsxWriter
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")
        self.generated_df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

    

        writer.close()

        messagebox.showinfo("Done", f"Saved to:\n{output_path}")



# Run App
root = tk.Tk()
root.state('zoomed')
app = ExcelCombinerApp(root)
root.mainloop()
