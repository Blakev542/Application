import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tkinter import Tk, filedialog
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter 
import customtkinter as ctk
import CTkListbox as ctklb
import requests, sys, os
import subprocess



#VERSION 1.0.1 WORKING
# UPDATED UI TO BE STYLIZED, FIXED SEARCH. 

class ExcelCombinerApp:
    def __init__(self, root):
        
        self.root = root
        self.root.title("Excel Column Combiner")
        self.files = []
        self.allcolumns = []
        self.columns = []
        self.selected = []
        self.df = pd.DataFrame()
        self.generated_df = pd.DataFrame()
        self.ALIASES = {
            "eqid": "eq_id",
            "student_id": "eq_id",
            "studentid": "eq_id",
        }
        self.view = []
        # UI Layout
        self.build_ui(

        ) 
        self.hover_index = None
        self.inlistbox = False

    
    
    # ________________ AUTO UPDATE _________________ #
    def auto_update(self):
        latest_url = "https://github.com/Blakev542/Application/releases/latest/download/ExcelCombiner.exe"
        exe_path = sys.executable
        temp_path = exe_path.replace(".exe", "_new.exe")
        updater_path = exe_path.replace(".exe", "_updater.bat")

        try:
            r = requests.get(latest_url, stream=True)
            if r.status_code == 200:
                with open(temp_path, "wb") as f:
                    for chunk in r.iter_content(1024 * 1024):
                        f.write(chunk)

                # Create updater batch file
                with open(updater_path, "w") as f:
                    f.write(f"""
    @echo off
    timeout /t 2 /nobreak >nul
    move /y "{temp_path}" "{exe_path}"
    start "" "{exe_path}"
    del "%~f0"
    """)

                subprocess.Popen(["cmd", "/c", updater_path], shell=True)
                sys.exit()

        except Exception as e:
            print("Update failed:", e)
            # ---------------- UI ----------------

    def build_ui(self):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                        background="#2b2b2b",
                        foreground="#FFFFFF",
                        fieldbackground="#2b2b2b",
                        rowheight=25
                        )
                
        style.configure("Treeview",
            font=("Segoe UI", 14),       
            rowheight=24                  
        )
        

        style.configure("Treeview.Heading",
                        background="#2b2b2b",
                        foreground="#FFFFFF",
                        font=("Segoe UI", 16, "bold"))
        # Main container
        self.frame = ctk.CTkFrame(self.root)
        self.frame.pack(fill="both", expand=True, padx=10, pady=10)

        # ---------- LEFT PANEL ----------
        left_panel = ctk.CTkFrame(self.frame, width=300)
        left_panel.pack(side="left", fill="y", padx=10, pady=10)
        left_panel.pack_propagate(False)
        ctk.CTkLabel(left_panel, text="📂 Selected Excel Files", font=("Segoe UI", 16, "bold")).pack(anchor="w")

        self.file_listbox = ctklb.CTkListbox(left_panel, height=160)
        self.file_listbox.pack(fill="x", pady=5)

        ctk.CTkButton(left_panel, text="Add Excel Files", command=self.add_files).pack(fill="x", pady=3)
        ctk.CTkButton(left_panel, text="Load Columns", command=self.load_columns).pack(fill="x", pady=3)
        ctk.CTkButton(left_panel, text="Clear Files", command=self.clear_files).pack(fill="x", pady=3)
        
        # Column Search
        ctk.CTkLabel(left_panel, text="🔍 Search Columns",font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(10,0))
        self.column_search_bar = ctk.CTkEntry(left_panel, placeholder_text="Type to filter...")
        self.column_search_bar.pack(fill="x", pady=5)
        self.column_search_bar.bind("<KeyRelease>", lambda e: self.show_columns_listbox())

       
        # COLUMN LIST BOX __________________________________________________________________

        self.column_listbox = tk.Listbox(left_panel,selectmode="multiple", height=90,           
            bg="#444348",        # background
            fg="white",          # text color
            font=("Segoe UI", 16),
            selectbackground="#a51f1f",   # highlight color
            selectforeground="black",     # text color when selected
            activestyle="none")
        self.column_listbox.pack(fill="both", pady=5,)
        self.column_listbox.bind("<Motion>", self.on_motion)
        self.column_listbox.bind("<Leave>", self.on_leave)
        self.column_listbox.bind("<Enter>", self.on_entry)

        # ____________________________________________________________________________________
   
        

      
        right_panel = ctk.CTkFrame(self.frame)
        right_panel.pack(side="right", fill="both", padx=10, pady=10 ,expand= True)

        ctk.CTkLabel(right_panel, text="Data Preview", font=("Segoe UI", 16, "bold")).pack(anchor="w")

        # Treeview (ttk still needed)
        self.tree = ttk.Treeview(right_panel)
        self.tree.pack(fill="both", expand=True)

        # Bottom buttons
        bottom_bar = ctk.CTkFrame(right_panel)
        bottom_bar.pack(fill="x", pady=5)
        ctk.CTkButton(bottom_bar, text="Add Column to View", command=self.show_columns).pack(side = "left", pady=3, padx=5)
        ctk.CTkButton(bottom_bar, text="Combine & Export", fg_color="green", command=self.combine_and_export).pack(side="right", padx=5)







    #----------------- HOVER handling --------------- #
    def on_motion(self, event):
        if not self.inlistbox:
            return 
        index = self.column_listbox.nearest(event.y)
        if self.hover_index == index:
            return

        if self.hover_index is not None:
            self.column_listbox.itemconfig(self.hover_index, bg="#333333")

        
        self.column_listbox.itemconfig(index, bg="#144870")
        self.hover_index = index
    
    def on_leave(self, event):
        self.inlistbox = False
        self.hover_index
        if self.hover_index is not None:
            self.column_listbox.itemconfig(self.hover_index, bg="#333333")
        self.hover_index = None
    def on_entry(self, event):
        self.inlistbox = True
    # ---------------- File Handling ----------------
    
    def add_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx *.xls")])
        for f in files:
            if f not in self.files:
                self.files.append(f)
                f = f.rsplit("/", 1)[-1]
                self.file_listbox.insert(tk.END, f)
      
    def clear_cache(self):
        for col in self.tree.get_children():
            self.tree.delete(col)
        self.columns = []
        self.allcolumns.clear()

                    
                    
    def normalize_key(self,key):
        key = self.ALIASES.get(key.replace(" ", "_").lower(), key.replace(" ", "_").lower())
        return key 
   
    def find_best_key(self, df, candidates):
        df_candidates = []
        df_final_candidates = []
        for c in df.columns:
            if df[c].nunique() > len(df) * 0.9:  # mostly unique
                
                df_candidates.append(c)
        for candidate in df_candidates:
            if candidate in candidates:
                df_final_candidates.append(candidate)   
            
        return df_final_candidates
    def sort_candidates_by_uniqueness(self, df, candidates):
        scored = []

        for c in candidates:
            if c in df.columns:
                score = df[c].nunique(dropna=True)
                scored.append((c, score))

        # highest uniqueness first
        scored.sort(key=lambda x: x[1], reverse=True)

        return [c for c, _ in scored]
    def find_keys(self):
        default_key = "eq_id"
        columns = []
        countmax = 0
        candidates = []
        for column in self.allcolumns:
            columns.append(self.normalize_key(column))
        for i in columns:
            if columns.count(i) > 1:
                candidates.append(i)
        for i in candidates:
            if candidates.count(i)>countmax:
                countmax = candidates.count(i)
                default_key = i
        candidates = list(set(candidates))
        self.primary_key = default_key
        return default_key, candidates

    def build_df(self):
        default_key = "eq_id"
        filesnotmerged = self.files.copy()
        default_key, candidates = self.find_keys()
        maxloops = 20
        loops = 0
        self.df = pd.DataFrame()

        while(len(filesnotmerged) > 0):
            loops += 1
     
            if loops >= maxloops:
                print("exceeded maximum loops")
                return self.df
                
            for file in filesnotmerged[:]:
                filename = file.rsplit("/", 1)[-1]
                df = pd.read_excel(file)
                df = self.clean_file(df)
                df = df.rename(columns=lambda c: self.normalize_key(c))
                
                df_candidates = []
                for i in candidates:
                    if i in self.find_best_key(df, candidates):
                        df_candidates.append(i)

                if default_key not in df_candidates and len(df_candidates) != 0:
             
                    default_key = df_candidates[0]
                #if len(df_candidates) == 0:
                #    df = self.aggregate_to_lists(df,default_key)

                if self.df.empty and default_key in df.columns:

                    self.df = df

                    if file in filesnotmerged:filesnotmerged.remove(file)
                    continue
                elif default_key in self.df.columns and default_key in df.columns and default_key in df_candidates:

                    self.df = df.merge(self.df, how="outer", on=default_key, suffixes=("_"+ filename,""))
                    
                    if file in filesnotmerged:filesnotmerged.remove(file)
                    continue
                else:
                    candidates = self.sort_candidates_by_uniqueness(df, candidates)
                    for candidate_key in candidates:
                        if candidate_key in self.df.columns and candidate_key in df.columns:
                            #if candidate_key not in df_candidates:
                                #df = self.aggregate_to_lists(df,candidate_key)
                            print("\ndefault key" , candidate_key,"\n\n\n\n\n")
                           
                            self.df = df.merge(self.df, how="outer", on=candidate_key, suffixes=("_"+ filename,""))

                            if file in filesnotmerged:filesnotmerged.remove(file)
                            break  
        self.columns = self.df.columns
        return self.df     
    def show_columns(self):
        self.selected = [self.column_listbox.get(i) for i in self.column_listbox.curselection()]
        for col in self.tree.get_children():
            self.tree.delete(col)
        self.generated_df = self.df.copy()
        self.view.extend(self.selected)
        self.generated_df.drop(columns=[c for c in self.df.columns if c not in self.view], inplace=True)
        self.generated_df = self.generated_df.drop_duplicates()
        self.generated_df = self.generated_df[self.view]
        self.show_columns_listbox()
        self.tree["columns"] = self.view
        self.tree["show"] = "headings"
     
        for col in self.view:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, stretch=False)
        for _, row in self.generated_df.iterrows():
            self.tree.insert("", tk.END, values=row.tolist())
                            
    def show_columns_listbox(self):
        self.column_listbox.pack_forget()
        self.column_listbox.delete(0, tk.END)
        self.columns = [c for c in self.columns if c not in self.view]
        string = self.column_search_bar.get()
        if string is not None and string != "":
            string = string.lower()
            columns = [c for c in self.columns if string in c.lower()]
        else:
            columns = sorted(self.columns)

        
        for col in columns:
            self.column_listbox.insert(tk.END,col,)
        self.column_listbox.pack(fill="both", pady=5)



    def clear_files(self):
        self.files.clear()
        self.file_listbox.delete(0, tk.END)
        self.column_listbox.delete(0, tk.END)
        self.columns.clear()

    # ---------------- Column Loading ----------------
    def load_columns(self):
        self.clear_cache()
        
        if not self.files:
            messagebox.showerror("Error", "No files selected")
            return
        
        
        
        for file in self.files:
            df = pd.read_excel(file)
     
            df = self.clean_file(df)
         
            for col in df.columns:
                col = col.replace(" ", "_").lower()
                self.allcolumns.append(str(col))
  
        
        self.build_df()
        self.show_columns_listbox()


    def clean_file(self, df):   

        
        first_clean_index = 0
        for i, row in df.iterrows():
            if not any(str(cell).startswith("Unnamed") for cell in row) and not any(pd.isna(cell) for cell in row):
                first_clean_index = i
                if i > 10:
                    return df
                df.columns = df.iloc[first_clean_index]



                df = df.iloc[first_clean_index+1:].reset_index(drop=True)

                return df
        
        return df 
            
           

    # ---------------- Combine + Export ----------------
    def combine_and_export(self):
        if self.generated_df.empty:
            messagebox.showerror("Error", "No data to export")
            return
        #selected = [self.column_listbox.get(i) for i in self.column_listbox.curselection()]
        
        #if not selected:
        #   messagebox.showerror("Error", "No columns selected")
        #   return
        output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialfile="output.xlsx",
        title="Save Excel File As"
        )
        
        if not output_path:
            return
        
      

        # Write with XlsxWriter
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")
        self.generated_df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

    

        writer.close()

        messagebox.showinfo("Done", f"Saved to:\n{output_path}")



# Run App

ctk.set_appearance_mode("dark")   
ctk.set_default_color_theme("blue")

root = ctk.CTk()

app = ExcelCombinerApp(root)
root.update()
root.state("zoomed")
root.after(3000, app.auto_update)
root.mainloop()

